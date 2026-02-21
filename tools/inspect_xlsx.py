import json
from openpyxl import load_workbook
wb = load_workbook('aa_test.xlsx', read_only=True, data_only=True)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
headers = rows[0] if rows else []
first_rows = rows[1:6]
print(json.dumps({'headers': headers, 'first_rows': first_rows}, default=str, ensure_ascii=False))
