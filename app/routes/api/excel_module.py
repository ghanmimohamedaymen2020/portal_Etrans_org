from datetime import datetime
import re

from flask import current_app, jsonify, request, Response
from flask_login import current_user, login_required
from sqlalchemy.exc import IntegrityError

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None

from app import db
from app.models.excel_module import (
    DEFAULT_EXCEL_COLUMNS,
    ExcelColumn,
    ExcelRecord,
    ExcelRecordExtraValue,
)
from app.routes.api import api_bp
from app.services.permission_service import has_permission


def _normalize_key(raw: str) -> str:
    value = (raw or "").strip().lower()
    value = re.sub(r"[^a-z0-9_]+", "_", value)
    value = re.sub(r"_+", "_", value).strip("_")
    return value or "colonne"


def _next_column_position() -> int:
    return (db.session.query(db.func.max(ExcelColumn.position)).scalar() or 0) + 1


def _parse_date(value):
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_number(value):
    if value is None:
        return None
    s = str(value).strip().replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _check_perm(code: str):
    if not current_user.is_authenticated:
        return jsonify({"message": "Non authentifié"}), 401
    if not has_permission(current_user, code):
        return jsonify({"message": "Accès refusé", "permission": code}), 403
    return None


def _normalize_excel_table_names(users_engine):
    """Ensure legacy table names exist even if a previous run renamed them."""
    if users_engine.dialect.name != "mssql":
        return

    rename_pairs = [
        ("suivie_cs_imp_columns", "excel_columns"),
        ("suivie_cs_imp", "excel_records"),
        ("suivie_cs_imp_extra_values", "excel_record_extra_values"),
    ]

    with users_engine.begin() as conn:
        existing_tables = {
            row[0]
            for row in conn.execute(
                db.text("SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA = 'dbo'")
            )
        }

        for source_name, target_name in rename_pairs:
            if target_name in existing_tables:
                continue
            if source_name not in existing_tables:
                continue
            conn.execute(db.text(f"EXEC sp_rename '[dbo].[{source_name}]', '{target_name}'"))
            existing_tables.remove(source_name)
            existing_tables.add(target_name)


def _ensure_module_ready():
    users_engine = db.engines.get("users")
    _normalize_excel_table_names(users_engine)
    ExcelColumn.__table__.create(bind=users_engine, checkfirst=True)
    ExcelRecord.__table__.create(bind=users_engine, checkfirst=True)
    ExcelRecordExtraValue.__table__.create(bind=users_engine, checkfirst=True)

    existing = {c.key: c for c in ExcelColumn.query.all()}
    changed = False
    for idx, (key, label, data_type, is_default) in enumerate(DEFAULT_EXCEL_COLUMNS):
        col = existing.get(key)
        if not col:
            db.session.add(
                ExcelColumn(
                    key=key,
                    label=label,
                    data_type=data_type,
                    is_default=is_default,
                    is_active=True,
                    position=idx,
                    created_by="system",
                )
            )
            changed = True
            continue
        if col.label != label:
            col.label = label
            changed = True
        if col.data_type != data_type:
            col.data_type = data_type
            changed = True
        if not col.is_default:
            col.is_default = True
            changed = True
        if col.position != idx:
            col.position = idx
            changed = True
    if changed:
        db.session.commit()


def _active_columns():
    return ExcelColumn.query.filter_by(is_active=True).order_by(ExcelColumn.position.asc(), ExcelColumn.id.asc()).all()


def _record_to_dict(record: ExcelRecord, columns):
    out = {
        "id": record.id,
        "created_at": record.created_at.isoformat() if record.created_at else None,
        "updated_at": record.updated_at.isoformat() if record.updated_at else None,
        "created_by": record.created_by,
        "updated_by": record.updated_by,
    }
    for key, _label, data_type, _is_default in DEFAULT_EXCEL_COLUMNS:
        value = getattr(record, key)
        if data_type == "date" and value is not None:
            out[key] = value.isoformat()
        else:
            out[key] = value

    extras = {ev.column_id: ev.value for ev in record.extra_values}
    for col in columns:
        if col.is_default:
            continue
        out[col.key] = extras.get(col.id)

    return out


def _apply_payload_to_record(record: ExcelRecord, payload: dict, columns):
    for key, _label, data_type, _is_default in DEFAULT_EXCEL_COLUMNS:
        raw = payload.get(key)
        if data_type == "date":
            setattr(record, key, _parse_date(raw))
        elif data_type == "number":
            setattr(record, key, _parse_number(raw))
        else:
            setattr(record, key, (str(raw).strip() if raw is not None else None) or None)

    by_key = {c.key: c for c in columns if not c.is_default}
    provided = payload.get("extra", {}) if isinstance(payload.get("extra", {}), dict) else {}

    existing = {ev.column_id: ev for ev in record.extra_values}
    for key, raw in provided.items():
        col = by_key.get(key)
        if not col:
            continue
        txt = (str(raw).strip() if raw is not None else "")
        item = existing.get(col.id)
        if not txt:
            if item:
                db.session.delete(item)
            continue
        if item:
            item.value = txt
        else:
            db.session.add(ExcelRecordExtraValue(record=record, column_id=col.id, value=txt))


@api_bp.route("/excel-module/columns", methods=["GET"])
@login_required
def excel_columns_list():
    err = _check_perm("excel.view")
    if err:
        return err
    _ensure_module_ready()
    cols = _active_columns()
    return jsonify([
        {
            "id": c.id,
            "key": c.key,
            "label": c.label,
            "data_type": c.data_type,
            "is_default": c.is_default,
            "is_active": c.is_active,
            "position": c.position,
        }
        for c in cols
    ])


@api_bp.route("/excel-module/columns", methods=["POST"])
@login_required
def excel_columns_create():
    err = _check_perm("excel.columns.manage")
    if err:
        return err
    _ensure_module_ready()

    payload = request.get_json() or {}
    label = (payload.get("label") or "").strip()
    if not label:
        return jsonify({"message": "Le libellé est obligatoire"}), 400

    key = _normalize_key(payload.get("key") or label)
    if ExcelColumn.query.filter_by(key=key).first():
        return jsonify({"message": "Cette colonne existe déjà"}), 409

    data_type = (payload.get("data_type") or "text").strip().lower()
    if data_type not in {"text", "number", "date"}:
        data_type = "text"

    max_pos = db.session.query(db.func.max(ExcelColumn.position)).scalar() or 0
    col = ExcelColumn(
        key=key,
        label=label,
        data_type=data_type,
        is_default=False,
        is_active=True,
        position=max_pos + 1,
        created_by=getattr(current_user, "username", None),
    )
    db.session.add(col)
    db.session.commit()
    return jsonify({"message": "Colonne ajoutée", "id": col.id, "key": col.key}), 201


@api_bp.route("/excel-module/columns/<int:column_id>", methods=["DELETE"])
@login_required
def excel_columns_delete(column_id: int):
    err = _check_perm("excel.columns.manage")
    if err:
        return err
    _ensure_module_ready()

    col = ExcelColumn.query.get_or_404(column_id)
    if col.is_default:
        return jsonify({"message": "Impossible de supprimer une colonne par défaut"}), 403

    ExcelRecordExtraValue.query.filter_by(column_id=col.id).delete()
    db.session.delete(col)
    db.session.commit()
    return jsonify({"message": "Colonne supprimée"})


@api_bp.route("/excel-module/records", methods=["GET"])
@login_required
def excel_records_list():
    err = _check_perm("excel.view")
    if err:
        return err
    _ensure_module_ready()

    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 100, type=int)
    # 0 means "all" — cap at 10 000 to avoid memory issues
    if per_page <= 0:
        per_page = 10000
    per_page = max(1, min(per_page, 10000))
    q = (request.args.get("q") or "").strip().lower()

    query = ExcelRecord.query
    if q:
        query = query.filter(
            db.or_(
                db.func.lower(db.func.coalesce(ExcelRecord.ref, "")).contains(q),
                db.func.lower(db.func.coalesce(ExcelRecord.shipper, "")).contains(q),
                db.func.lower(db.func.coalesce(ExcelRecord.forwarder, "")).contains(q),
                db.func.lower(db.func.coalesce(ExcelRecord.offer_ref, "")).contains(q),
                db.func.lower(db.func.coalesce(ExcelRecord.cgnee, "")).contains(q),
            )
        )

    pag = query.order_by(ExcelRecord.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    cols = _active_columns()

    return jsonify({
        "items": [_record_to_dict(r, cols) for r in pag.items],
        "total": pag.total,
        "pages": pag.pages,
        "current_page": pag.page,
    })


@api_bp.route("/excel-module/records", methods=["POST"])
@login_required
def excel_records_create():
    err = _check_perm("excel.create")
    if err:
        return err
    _ensure_module_ready()

    payload = request.get_json() or {}
    cols = _active_columns()
    rec = ExcelRecord(created_by=getattr(current_user, "username", None), updated_by=getattr(current_user, "username", None))
    _apply_payload_to_record(rec, payload, cols)
    db.session.add(rec)
    db.session.commit()

    return jsonify({"message": "Enregistrement créé", "id": rec.id}), 201


@api_bp.route("/excel-module/records/<int:record_id>", methods=["PUT"])
@login_required
def excel_records_update(record_id: int):
    err = _check_perm("excel.edit")
    if err:
        return err
    _ensure_module_ready()

    payload = request.get_json() or {}
    cols = _active_columns()
    rec = ExcelRecord.query.get_or_404(record_id)
    _apply_payload_to_record(rec, payload, cols)
    rec.updated_by = getattr(current_user, "username", None)
    db.session.commit()

    return jsonify({"message": "Enregistrement mis à jour"})


@api_bp.route("/excel-module/records/<int:record_id>", methods=["DELETE"])
@login_required
def excel_records_delete(record_id: int):
    err = _check_perm("excel.delete")
    if err:
        return err
    _ensure_module_ready()

    rec = ExcelRecord.query.get_or_404(record_id)
    db.session.delete(rec)
    db.session.commit()
    return jsonify({"message": "Enregistrement supprimé"})


@api_bp.route("/excel-module/records/export.xlsx", methods=["GET"])
@login_required
def excel_records_export_xlsx():
    err = _check_perm("excel.export.xlsx")
    if err:
        return err
    _ensure_module_ready()

    if Workbook is None:
        return jsonify({"message": "openpyxl non disponible"}), 500

    q = (request.args.get("q") or "").strip().lower()
    query = ExcelRecord.query
    if q:
        query = query.filter(
            db.or_(
                db.func.lower(db.func.coalesce(ExcelRecord.ref, "")).contains(q),
                db.func.lower(db.func.coalesce(ExcelRecord.shipper, "")).contains(q),
                db.func.lower(db.func.coalesce(ExcelRecord.forwarder, "")).contains(q),
                db.func.lower(db.func.coalesce(ExcelRecord.offer_ref, "")).contains(q),
                db.func.lower(db.func.coalesce(ExcelRecord.cgnee, "")).contains(q),
            )
        )
    rows = query.order_by(ExcelRecord.created_at.desc()).all()

    cols = _active_columns()
    wb = Workbook()
    ws = wb.active
    ws.title = "Suivie CS IMP"

    headers = [c.label for c in cols]
    ws.append(headers)

    for rec in rows:
        data = _record_to_dict(rec, cols)
        ws.append([data.get(c.key) for c in cols])

    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
            if col_letter is None:
                col_letter = cell.column_letter
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 50)

    import io
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return Response(
        out.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="suivie_cs_imp_export.xlsx"'},
    )


@api_bp.route("/excel-module/import.xlsx", methods=["POST"])
@login_required
def excel_records_import_xlsx():
    err = _check_perm("excel.create")
    if err:
        return err

    try:
        _ensure_module_ready()

        if load_workbook is None:
            return jsonify({"message": "openpyxl non disponible"}), 500

        uploaded = request.files.get("file")
        if not uploaded or not uploaded.filename:
            return jsonify({"message": "Fichier Excel requis"}), 400

        filename = uploaded.filename.lower()
        if not (filename.endswith(".xlsx") or filename.endswith(".xlsm")):
            return jsonify({"message": "Format non supporté. Utilisez .xlsx"}), 400

        wb = load_workbook(uploaded, data_only=True)
        ws = wb.active

        header_row = None
        header_index = None
        scan_max = min(ws.max_row or 1, 25)
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=scan_max, values_only=True), start=1):
            if not row:
                continue
            candidate = [str(h).strip() if h is not None else "" for h in row]
            non_empty = [h for h in candidate if h]
            if len(non_empty) >= 2:
                header_row = candidate
                header_index = idx
                break

        if not header_row or not header_index:
            return jsonify({"message": "Aucune entête détectée"}), 400

        headers = header_row

        columns = _active_columns()
        by_normalized = {}
        for c in columns:
            by_normalized[_normalize_key(c.key)] = c
            by_normalized[_normalize_key(c.label)] = c

        auto_create_allowed = has_permission(current_user, "excel.columns.manage")
        unknown_headers = []

        for h in headers:
            if not h:
                continue
            n = _normalize_key(h)
            if n in by_normalized:
                continue
            if auto_create_allowed:
                key = n
                suffix = 1
                while ExcelColumn.query.filter_by(key=key).first() is not None:
                    key = f"{n}_{suffix}"
                    suffix += 1
                col = ExcelColumn(
                    key=key,
                    label=h,
                    data_type="text",
                    is_default=False,
                    is_active=True,
                    position=_next_column_position(),
                    created_by=getattr(current_user, "username", None),
                )
                db.session.add(col)
                db.session.flush()
                by_normalized[_normalize_key(col.key)] = col
                by_normalized[_normalize_key(col.label)] = col
            else:
                unknown_headers.append(h)

        columns = _active_columns()
        imported = 0
        skipped = 0
        duplicate_refs = []

        existing_refs = {
            r for (r,) in db.session.query(ExcelRecord.ref).filter(ExcelRecord.ref.isnot(None)).all()
        }

        for row_number, row in enumerate(
            ws.iter_rows(min_row=header_index + 1, values_only=True),
            start=header_index + 1,
        ):
            if not row or not any(v is not None and str(v).strip() != "" for v in row):
                continue

            payload = {"extra": {}}
            for idx, cell in enumerate(row):
                if idx >= len(headers):
                    continue
                header = headers[idx]
                if not header:
                    continue
                col = by_normalized.get(_normalize_key(header))
                if not col:
                    continue
                value = "" if cell is None else str(cell)
                if col.is_default:
                    payload[col.key] = value
                else:
                    payload["extra"][col.key] = value

            row_ref = (payload.get("ref") or "").strip()
            if row_ref and row_ref in existing_refs:
                skipped += 1
                duplicate_refs.append({"row": row_number, "ref": row_ref})
                continue

            rec = ExcelRecord(
                created_by=getattr(current_user, "username", None),
                updated_by=getattr(current_user, "username", None),
            )
            _apply_payload_to_record(rec, payload, columns)
            db.session.add(rec)
            if row_ref:
                existing_refs.add(row_ref)
            imported += 1

        db.session.commit()
        return jsonify({
            "message": "Import terminé",
            "imported": imported,
            "skipped": skipped,
            "duplicate_refs": duplicate_refs,
            "ignored_headers": unknown_headers,
        })
    except IntegrityError:
        db.session.rollback()
        current_app.logger.exception("Erreur d'integrite SQL pendant import XLSX")
        return jsonify({
            "message": "Import impossible: conflit de donnees en base (contrainte SQL). "
                       "Verifiez les doublons et la coherence des colonnes puis reessayez.",
        }), 500
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Erreur import XLSX GESTION DES ROUTING TGY TUNISIE")
        return jsonify({"message": f"Erreur import XLSX: {exc.__class__.__name__}"}), 500
